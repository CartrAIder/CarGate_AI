# -*- coding: utf-8 -*-
"""Backend hand-off package: product master (xlsx) + one front-facing image per
product, filenames paired to the sheet.

  service_products/
    products.xlsx           바코드 · 상품명 · 카테고리 · 이미지파일 ...
    products.csv            (동일 내용 CSV 백업)
    images/<sku>__<한글명>.jpg   상품별 정면샷 1장 (Excel '이미지파일'과 파일명 일치)
  service_products.zip       위 폴더 압축본 (팀원 전달용)

  conda run -n cartgate python build_handoff.py
"""
import csv, shutil, glob, zipfile, sys
from pathlib import Path
import cv2, numpy as np
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SRC = Path("dataset/images")
OUT = Path("service_products")
IMGDIR = OUT / "images"
IMG_EXT = {".jpg", ".jpeg", ".png", ".webp"}
FONTP = sorted(glob.glob("/usr/share/fonts/truetype/nanum/NanumGothic*.ttf"))

KOR = {
    "S0001": "알로에", "S0002": "사과", "S0003": "가방", "S0004": "밴드", "S0005": "건전지(AA)",
    "S0006": "시계", "S0007": "콜라", "S0008": "큐브", "S0009": "컵", "S0010": "인형",
    "S0011": "공학용계산기", "S0012": "칼", "S0013": "뒤집개", "S0014": "게임패드",
    "S0015": "악력기", "S0016": "기타", "S0017": "핸드크림", "S0018": "키보드", "S0019": "립밤",
    "S0020": "모니터", "S0021": "마우스", "S0022": "마우스패드", "S0023": "접시", "S0024": "보조배터리",
    "S0025": "가위", "S0026": "각티슈", "S0027": "치약", "S0028": "우산", "S0029": "USB메모리",
    "S0030": "지갑", "S0031": "공기청정기", "S0032": "김", "S0033": "노세범", "S0034": "독서받침대",
    "S0035": "마스크팩", "S0036": "물티슈", "S0037": "배홍동칼빔면", "S0038": "세럼", "S0039": "수분크림",
    "S0040": "썬스틱", "S0041": "오뚜기작은밥", "S0042": "요가매트", "S0043": "운동화", "S0044": "주방세제",
    "S0045": "참치캔", "S0046": "체중계", "S0047": "쿠션", "S0048": "크록스", "S0049": "텀블러",
    "S0050": "통조림닭가슴살", "S0051": "휴지",
}
CAT = {
    "식품": ["S0002", "S0032", "S0037", "S0041", "S0045", "S0050"],
    "음료": ["S0001", "S0007"],
    "화장품/미용": ["S0017", "S0019", "S0033", "S0035", "S0038", "S0039", "S0040", "S0047"],
    "생활용품": ["S0004", "S0006", "S0025", "S0026", "S0027", "S0028", "S0034", "S0036", "S0044", "S0051"],
    "디지털/가전": ["S0005", "S0011", "S0014", "S0018", "S0020", "S0021", "S0022", "S0024", "S0029", "S0031", "S0046"],
    "주방용품": ["S0009", "S0013", "S0023", "S0049"],
    "패션잡화": ["S0003", "S0030", "S0043", "S0048"],
    "완구/취미": ["S0008", "S0010", "S0012", "S0016"],
    "스포츠/레저": ["S0015", "S0042"],
}
CAT_OF = {s: c for c, ss in CAT.items() for s in ss}
FRONT = {"S0001": 1, "S0008": 1, "S0018": 1, "S0021": 2, "S0031": 1, "S0037": 2,
         "S0038": 1, "S0039": 1, "S0040": 1, "S0041": 2, "S0043": 1}   # else index 0


def load_products():
    m = {}
    for r in csv.DictReader(open("products.csv", encoding="utf-8-sig")):
        m[r["sku_id"]] = {"en": r["name"], "barcode": r.get("barcode", "").strip()}
    return m


def main():
    no_images = "--no-images" in sys.argv          # front shots excluded (user adds their own)
    imgcol = "이미지파일(권장명)" if no_images else "이미지파일"
    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True)
    if not no_images:
        IMGDIR.mkdir(parents=True)
    prod = load_products()
    dirs = {p.name.split("__")[0]: p for p in SRC.iterdir() if p.is_dir()}
    rows = []
    for sku in sorted(dirs):
        imgs = [p for p in sorted(dirs[sku].iterdir()) if p.suffix.lower() in IMG_EXT]
        if not imgs:
            continue
        kor = KOR.get(sku, prod.get(sku, {}).get("en", sku))
        base = f"{sku}__{kor.replace(' ', '_')}"
        if no_images:
            imgval = f"{base}.jpg"                  # 권장 파일명: 나중에 이 이름으로 넣으면 행과 짝맞음
        else:
            src = imgs[min(FRONT.get(sku, 0), len(imgs) - 1)]
            fname = f"{base}{src.suffix.lower()}"
            shutil.copy2(src, IMGDIR / fname)
            imgval = f"images/{fname}"
        rows.append({
            "sku_id": sku,
            "바코드(EAN13)": prod.get(sku, {}).get("barcode", ""),
            "상품명": kor,
            "영문명": prod.get(sku, {}).get("en", ""),
            "카테고리": CAT_OF.get(sku, "미분류"),
            imgcol: imgval,
        })

    cols = ["sku_id", "바코드(EAN13)", "상품명", "영문명", "카테고리", imgcol]
    with open(OUT / "products.csv", "w", encoding="utf-8-sig", newline="\n") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(rows)

    wb = Workbook(); ws = wb.active; ws.title = "상품마스터"
    hd_fill = PatternFill("solid", fgColor="1F2937"); hd_font = Font(color="FFFFFF", bold=True)
    ws.append(cols)
    for c in ws[1]:
        c.fill = hd_fill; c.font = hd_font; c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r[c] for c in cols])
    # force barcode as text so leading zeros survive
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for c in row:
            c.number_format = "@"
    widths = [10, 18, 16, 22, 14, 34]
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w_
    ws.freeze_panes = "A2"
    wb.save(OUT / "products.xlsx")

    with zipfile.ZipFile("service_products.zip", "w", zipfile.ZIP_DEFLATED) as z:
        for p in sorted(OUT.rglob("*")):
            if p.is_file():
                z.write(p, p.relative_to(OUT.parent))

    no_bc = [r["sku_id"] + "/" + r["상품명"] for r in rows if not r["바코드(EAN13)"]]
    nimg = 0 if no_images else len(list(IMGDIR.iterdir()))
    print(f"products: {len(rows)}  images: {nimg}{' (정면샷 제외)' if no_images else ''}  "
          f"with-barcode: {sum(1 for r in rows if r['바코드(EAN13)'])}/{len(rows)}")
    print(f"categories: " + ", ".join(f"{c}:{len(s)}" for c, s in CAT.items()))
    print(f"NO barcode ({len(no_bc)}): " + ", ".join(no_bc))
    print(f"-> service_products/ ({'xlsx+csv' if no_images else 'xlsx+csv+images'}) + service_products.zip")

    if not no_images:
        preview(rows)   # verify chosen front shots


def preview(rows):
    cell, lab = 150, 40
    cols = 6
    n = len(rows); r_ = (n + cols - 1) // cols
    font = ImageFont.truetype(FONTP[0], 14) if FONTP else ImageFont.load_default()
    sfont = ImageFont.truetype(FONTP[0], 12) if FONTP else ImageFont.load_default()
    W, H = cols * (cell + 8) + 8, r_ * (cell + lab + 8) + 8
    canvas = Image.new("RGB", (W, H), (18, 22, 28)); d = ImageDraw.Draw(canvas)
    for i, r in enumerate(rows):
        cy, cx = divmod(i, cols)
        x, y = 8 + cx * (cell + 8), 8 + cy * (cell + lab + 8)
        b = cv2.imread(str(OUT / r["이미지파일"]))
        if b is not None:
            s = cell / max(b.shape[:2]); rz = cv2.resize(b, (int(b.shape[1] * s), int(b.shape[0] * s)))
            tile = np.full((cell, cell, 3), (245, 245, 245), np.uint8)
            yy, xx = (cell - rz.shape[0]) // 2, (cell - rz.shape[1]) // 2
            tile[yy:yy + rz.shape[0], xx:xx + rz.shape[1]] = rz
            canvas.paste(Image.fromarray(cv2.cvtColor(tile, cv2.COLOR_BGR2RGB)), (x, y))
        d.text((x + 2, y + cell + 2), f"{r['sku_id']} {r['상품명']}", fill=(255, 230, 120), font=font)
        d.text((x + 2, y + cell + 20), f"{r['카테고리']} · {r['바코드(EAN13)'] or '바코드없음'}",
               fill=(150, 200, 240), font=sfont)
    canvas.save("out/handoff_preview.png")
    print("preview -> out/handoff_preview.png")


if __name__ == "__main__":
    main()
